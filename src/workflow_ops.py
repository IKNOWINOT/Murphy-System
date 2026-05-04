"""
workflow_ops.py — PATCH-180
WorkOps: Unified workflow operations center.

Provides:
- 9 common business workflow templates with explicit pickup/putdown points
- SQLite ledger for all workflow activity
- Excel export (openpyxl)
- REST API endpoints

Each workflow step has a strict handoff contract:
  pickup_agent picks it up → does work → putdown_agent passes it on
"""

import sqlite3
import json
import uuid
import logging
import os
from datetime import datetime, timezone
from typing import Dict, List, Optional, Any
from contextlib import contextmanager

logger = logging.getLogger("murphy.workflow_ops")

DB_PATH = "/var/lib/murphy-production/workflow_ops.db"

# ── 9 Common Workflow Templates ───────────────────────────────────────────────

WORKFLOW_TEMPLATES = [
    {
        "id": "tmpl_lead_to_customer",
        "name": "Lead to Customer",
        "category": "Sales",
        "description": "Full sales cycle from raw lead signal to signed customer",
        "icon": "💰",
        "priority_default": "high",
        "steps": [
            {
                "index": 0,
                "name": "Signal Collection",
                "pickup_agent": "collector",
                "putdown_agent": "collector",
                "description": "Collect and deduplicate incoming lead signals",
                "pickup_form": ["lead_source", "contact_email", "company_name", "signal_strength"],
                "putdown_form": ["enriched_profile", "intent_score", "dedup_hash", "next_action"],
                "expected_duration_m": 5
            },
            {
                "index": 1,
                "name": "Intent Translation",
                "pickup_agent": "translator",
                "putdown_agent": "translator",
                "description": "Translate lead signal into structured business intent",
                "pickup_form": ["enriched_profile", "intent_score"],
                "putdown_form": ["business_need", "budget_estimate", "urgency", "fit_score"],
                "expected_duration_m": 10
            },
            {
                "index": 2,
                "name": "Executive Qualification",
                "pickup_agent": "exec_admin",
                "putdown_agent": "exec_admin",
                "description": "ExecAdmin qualifies and assigns follow-up strategy",
                "pickup_form": ["business_need", "fit_score", "budget_estimate"],
                "putdown_form": ["qualification_verdict", "assigned_tier", "outreach_script", "follow_up_date"],
                "expected_duration_m": 15
            },
            {
                "index": 3,
                "name": "Outreach Execution",
                "pickup_agent": "executor",
                "putdown_agent": "executor",
                "description": "Execute outreach — email, Telegram notification, CRM entry",
                "pickup_form": ["outreach_script", "contact_email", "assigned_tier"],
                "putdown_form": ["email_sent_id", "crm_record_id", "response_received", "outcome"],
                "expected_duration_m": 5
            },
            {
                "index": 4,
                "name": "Audit & Close",
                "pickup_agent": "auditor",
                "putdown_agent": "auditor",
                "description": "Verify outreach compliance, log ROI, close workflow",
                "pickup_form": ["email_sent_id", "crm_record_id", "outcome"],
                "putdown_form": ["audit_verdict", "roi_logged", "compliance_flags", "final_status"],
                "expected_duration_m": 3
            }
        ]
    },
    {
        "id": "tmpl_content_publish",
        "name": "Content Publish",
        "category": "Marketing",
        "description": "From content idea to published, audited post",
        "icon": "📝",
        "priority_default": "medium",
        "steps": [
            {
                "index": 0,
                "name": "Brief Creation",
                "pickup_agent": "translator",
                "putdown_agent": "translator",
                "description": "Translate business goal into content brief",
                "pickup_form": ["topic", "target_audience", "business_goal", "channel"],
                "putdown_form": ["content_brief", "tone_guide", "keywords", "word_count_target"],
                "expected_duration_m": 10
            },
            {
                "index": 1,
                "name": "Content Production",
                "pickup_agent": "prod_ops",
                "putdown_agent": "prod_ops",
                "description": "Draft content according to brief",
                "pickup_form": ["content_brief", "tone_guide", "keywords"],
                "putdown_form": ["draft_content", "draft_version", "assets_urls", "ready_for_review"],
                "expected_duration_m": 30
            },
            {
                "index": 2,
                "name": "Publish",
                "pickup_agent": "executor",
                "putdown_agent": "executor",
                "description": "Publish to target channel(s)",
                "pickup_form": ["draft_content", "channel", "publish_time"],
                "putdown_form": ["published_url", "publish_timestamp", "platform_id"],
                "expected_duration_m": 5
            },
            {
                "index": 3,
                "name": "Compliance Check",
                "pickup_agent": "auditor",
                "putdown_agent": "auditor",
                "description": "Verify content compliance — GDPR, brand standards",
                "pickup_form": ["published_url", "platform_id"],
                "putdown_form": ["compliance_verdict", "flags", "roi_estimate"],
                "expected_duration_m": 5
            }
        ]
    },
    {
        "id": "tmpl_compliance_scan",
        "name": "Compliance Scan",
        "category": "Legal / Compliance",
        "description": "Full compliance scan across active frameworks",
        "icon": "🛡️",
        "priority_default": "high",
        "steps": [
            {
                "index": 0,
                "name": "Scan Trigger",
                "pickup_agent": "auditor",
                "putdown_agent": "auditor",
                "description": "Trigger full framework scan, collect findings",
                "pickup_form": ["trigger_reason", "frameworks_to_scan", "scope"],
                "putdown_form": ["findings_count", "critical_count", "scan_report_url"],
                "expected_duration_m": 15
            },
            {
                "index": 1,
                "name": "Soul Gate Review",
                "pickup_agent": "rosetta",
                "putdown_agent": "rosetta",
                "description": "Rosetta evaluates findings against North Star + HarmFloor",
                "pickup_form": ["findings_count", "critical_count", "scan_report_url"],
                "putdown_form": ["soul_verdict", "blocked_actions", "approved_remediations"],
                "expected_duration_m": 5
            },
            {
                "index": 2,
                "name": "HITL Escalation",
                "pickup_agent": "hitl",
                "putdown_agent": "hitl",
                "description": "Human-in-the-loop review of critical findings",
                "pickup_form": ["soul_verdict", "blocked_actions", "critical_count"],
                "putdown_form": ["human_decision", "approved_by", "notes"],
                "expected_duration_m": 60
            },
            {
                "index": 3,
                "name": "Remediation Dispatch",
                "pickup_agent": "exec_admin",
                "putdown_agent": "exec_admin",
                "description": "Dispatch remediation tasks and update compliance state",
                "pickup_form": ["human_decision", "approved_remediations"],
                "putdown_form": ["tasks_dispatched", "compliance_state_updated", "next_scan_date"],
                "expected_duration_m": 10
            }
        ]
    },
    {
        "id": "tmpl_revenue_driver",
        "name": "Revenue Driver",
        "category": "Finance",
        "description": "Cognitive executive drives the revenue cycle",
        "icon": "📈",
        "priority_default": "critical",
        "steps": [
            {
                "index": 0,
                "name": "Blocker Scan",
                "pickup_agent": "exec_admin",
                "putdown_agent": "exec_admin",
                "description": "Scan CRM, trials, onboarding, workflow failures for revenue blockers",
                "pickup_form": ["scan_scope", "time_horizon_days"],
                "putdown_form": ["blockers_found", "total_weight", "top_blocker_type"],
                "expected_duration_m": 5
            },
            {
                "index": 1,
                "name": "Directive Generation",
                "pickup_agent": "exec_admin",
                "putdown_agent": "exec_admin",
                "description": "Generate and dispatch revenue directives",
                "pickup_form": ["blockers_found", "top_blocker_type"],
                "putdown_form": ["directives_issued", "escalations", "automations_triggered"],
                "expected_duration_m": 5
            },
            {
                "index": 2,
                "name": "Execution",
                "pickup_agent": "executor",
                "putdown_agent": "executor",
                "description": "Execute directives — emails, CRM updates, Stripe actions",
                "pickup_form": ["directives_issued"],
                "putdown_form": ["actions_completed", "revenue_events_created", "failures"],
                "expected_duration_m": 10
            },
            {
                "index": 3,
                "name": "ROI Audit",
                "pickup_agent": "auditor",
                "putdown_agent": "auditor",
                "description": "Audit outcomes and log ROI",
                "pickup_form": ["actions_completed", "revenue_events_created"],
                "putdown_form": ["roi_delta", "audit_score", "recommendations"],
                "expected_duration_m": 5
            }
        ]
    },
    {
        "id": "tmpl_incident_response",
        "name": "Incident Response",
        "category": "Operations",
        "description": "Detect, contain, resolve, and document incidents",
        "icon": "🚨",
        "priority_default": "critical",
        "steps": [
            {
                "index": 0,
                "name": "Detection & Triage",
                "pickup_agent": "hitl",
                "putdown_agent": "hitl",
                "description": "Human-in-the-loop confirms incident and sets severity",
                "pickup_form": ["incident_type", "severity", "affected_system", "detected_at"],
                "putdown_form": ["severity_confirmed", "incident_id", "containment_needed", "human_notified"],
                "expected_duration_m": 15
            },
            {
                "index": 1,
                "name": "Compliance Assessment",
                "pickup_agent": "auditor",
                "putdown_agent": "auditor",
                "description": "Assess regulatory obligations (breach notification etc.)",
                "pickup_form": ["incident_type", "severity_confirmed", "affected_system"],
                "putdown_form": ["notification_required", "frameworks_triggered", "deadline"],
                "expected_duration_m": 30
            },
            {
                "index": 2,
                "name": "Containment & Ops",
                "pickup_agent": "prod_ops",
                "putdown_agent": "prod_ops",
                "description": "Contain the incident, apply fixes",
                "pickup_form": ["incident_type", "affected_system", "containment_needed"],
                "putdown_form": ["containment_actions", "system_restored", "root_cause"],
                "expected_duration_m": 60
            },
            {
                "index": 3,
                "name": "Resolution & Close",
                "pickup_agent": "executor",
                "putdown_agent": "executor",
                "description": "Execute notifications, update status page, close incident",
                "pickup_form": ["root_cause", "notification_required", "system_restored"],
                "putdown_form": ["notifications_sent", "status_page_updated", "postmortem_url"],
                "expected_duration_m": 20
            }
        ]
    },
    {
        "id": "tmpl_customer_onboarding",
        "name": "Customer Onboarding",
        "category": "Customer Success",
        "description": "New customer from signup to activated, fully onboarded",
        "icon": "🚀",
        "priority_default": "high",
        "steps": [
            {
                "index": 0,
                "name": "Signup Collection",
                "pickup_agent": "collector",
                "putdown_agent": "collector",
                "description": "Collect and validate new customer signup data",
                "pickup_form": ["customer_email", "tier", "signup_source", "company"],
                "putdown_form": ["account_id", "tier_confirmed", "validation_status"],
                "expected_duration_m": 2
            },
            {
                "index": 1,
                "name": "Onboarding Schedule",
                "pickup_agent": "scheduler",
                "putdown_agent": "scheduler",
                "description": "Schedule onboarding sequence and kickoff meeting",
                "pickup_form": ["account_id", "tier_confirmed"],
                "putdown_form": ["schedule_id", "kickoff_time", "onboarding_plan_url"],
                "expected_duration_m": 5
            },
            {
                "index": 2,
                "name": "Activation Execution",
                "pickup_agent": "executor",
                "putdown_agent": "executor",
                "description": "Send welcome email, provision account, create CRM record",
                "pickup_form": ["account_id", "customer_email", "tier_confirmed", "kickoff_time"],
                "putdown_form": ["welcome_email_id", "account_provisioned", "crm_record_id"],
                "expected_duration_m": 10
            },
            {
                "index": 3,
                "name": "Onboarding Audit",
                "pickup_agent": "auditor",
                "putdown_agent": "auditor",
                "description": "Verify activation, log ROI, confirm GDPR consent",
                "pickup_form": ["account_id", "welcome_email_id", "account_provisioned"],
                "putdown_form": ["activation_confirmed", "gdpr_consent_logged", "first_value_event"],
                "expected_duration_m": 5
            }
        ]
    },
    {
        "id": "tmpl_api_acquisition",
        "name": "API Acquisition",
        "category": "Engineering",
        "description": "Murphy self-activates and tests new APIs autonomously",
        "icon": "⚡",
        "priority_default": "medium",
        "steps": [
            {
                "index": 0,
                "name": "Discovery",
                "pickup_agent": "collector",
                "putdown_agent": "collector",
                "description": "Discover available free APIs matching current capability gaps",
                "pickup_form": ["capability_gap", "api_category", "max_results"],
                "putdown_form": ["candidates_found", "top_candidate", "auth_type"],
                "expected_duration_m": 10
            },
            {
                "index": 1,
                "name": "Acquisition & Test",
                "pickup_agent": "executor",
                "putdown_agent": "executor",
                "description": "Register, obtain key, test endpoint",
                "pickup_form": ["top_candidate", "auth_type"],
                "putdown_form": ["api_key_obtained", "test_result", "latency_ms", "capabilities"],
                "expected_duration_m": 15
            },
            {
                "index": 2,
                "name": "Integration Audit",
                "pickup_agent": "auditor",
                "putdown_agent": "auditor",
                "description": "Verify API doesn't introduce security or compliance risks",
                "pickup_form": ["top_candidate", "test_result", "capabilities"],
                "putdown_form": ["security_verdict", "compliance_flags", "approved_for_prod"],
                "expected_duration_m": 5
            }
        ]
    },
    {
        "id": "tmpl_morning_brief",
        "name": "Morning Brief",
        "category": "Management",
        "description": "Daily executive briefing — system health, revenue, priorities",
        "icon": "☀️",
        "priority_default": "medium",
        "steps": [
            {
                "index": 0,
                "name": "Signal Aggregation",
                "pickup_agent": "collector",
                "putdown_agent": "collector",
                "description": "Aggregate overnight signals from all sources",
                "pickup_form": ["time_window_hours", "signal_types"],
                "putdown_form": ["signals_collected", "notable_events", "anomalies"],
                "expected_duration_m": 5
            },
            {
                "index": 1,
                "name": "Executive Summary",
                "pickup_agent": "exec_admin",
                "putdown_agent": "exec_admin",
                "description": "Compose executive summary from signals",
                "pickup_form": ["signals_collected", "notable_events", "anomalies"],
                "putdown_form": ["summary_text", "action_items", "revenue_delta"],
                "expected_duration_m": 10
            },
            {
                "index": 2,
                "name": "Soul Review",
                "pickup_agent": "rosetta",
                "putdown_agent": "rosetta",
                "description": "Rosetta applies soul filter — align summary with North Star",
                "pickup_form": ["summary_text", "action_items"],
                "putdown_form": ["reviewed_summary", "soul_notes", "blocked_items"],
                "expected_duration_m": 3
            },
            {
                "index": 3,
                "name": "Delivery",
                "pickup_agent": "executor",
                "putdown_agent": "executor",
                "description": "Send brief via email to cpost + hpost",
                "pickup_form": ["reviewed_summary", "action_items"],
                "putdown_form": ["email_sent_id", "recipients", "delivered_at"],
                "expected_duration_m": 2
            }
        ]
    },
    {
        "id": "tmpl_self_patch",
        "name": "Self-Patch",
        "category": "Engineering",
        "description": "Murphy identifies a gap, writes a patch, reviews, deploys",
        "icon": "🔧",
        "priority_default": "medium",
        "steps": [
            {
                "index": 0,
                "name": "Gap Identification",
                "pickup_agent": "prod_ops",
                "putdown_agent": "prod_ops",
                "description": "MurphyMind identifies a gap or failure mode",
                "pickup_form": ["gap_description", "affected_module", "severity"],
                "putdown_form": ["gap_id", "root_cause", "proposed_fix_summary"],
                "expected_duration_m": 10
            },
            {
                "index": 1,
                "name": "Code Generation",
                "pickup_agent": "executor",
                "putdown_agent": "executor",
                "description": "ForgeEngine generates the patch code",
                "pickup_form": ["gap_id", "root_cause", "proposed_fix_summary"],
                "putdown_form": ["patch_code", "files_changed", "patch_id"],
                "expected_duration_m": 15
            },
            {
                "index": 2,
                "name": "Critic Gate",
                "pickup_agent": "auditor",
                "putdown_agent": "auditor",
                "description": "MurphyCritic reviews patch for 10 failure modes",
                "pickup_form": ["patch_code", "files_changed"],
                "putdown_form": ["critic_verdict", "failures_found", "approved_to_deploy"],
                "expected_duration_m": 5
            },
            {
                "index": 3,
                "name": "Deploy & Commission",
                "pickup_agent": "prod_ops",
                "putdown_agent": "prod_ops",
                "description": "Deploy patch, restart service, commission endpoints",
                "pickup_form": ["patch_id", "approved_to_deploy"],
                "putdown_form": ["deployed_at", "service_healthy", "endpoints_passing", "patch_number"],
                "expected_duration_m": 10
            }
        ]
    }
]

# ── DB Layer ───────────────────────────────────────────────────────────────────

def _init_db():
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.row_factory = sqlite3.Row
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS workflow_templates (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            category TEXT,
            description TEXT,
            icon TEXT,
            steps_json TEXT NOT NULL,
            created_at TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS workflow_instances (
            id TEXT PRIMARY KEY,
            template_id TEXT NOT NULL,
            name TEXT NOT NULL,
            category TEXT,
            status TEXT DEFAULT 'pending',
            priority TEXT DEFAULT 'medium',
            current_step INTEGER DEFAULT 0,
            total_steps INTEGER DEFAULT 0,
            started_at TEXT,
            completed_at TEXT,
            owner_agent TEXT,
            account_id TEXT,
            context_json TEXT DEFAULT '{}',
            created_at TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS workflow_steps (
            id TEXT PRIMARY KEY,
            instance_id TEXT NOT NULL,
            step_index INTEGER NOT NULL,
            step_name TEXT NOT NULL,
            pickup_agent TEXT,
            pickup_at TEXT,
            pickup_data TEXT DEFAULT '{}',
            putdown_agent TEXT,
            putdown_at TEXT,
            putdown_data TEXT DEFAULT '{}',
            handoff_notes TEXT,
            status TEXT DEFAULT 'pending',
            duration_s REAL,
            created_at TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS workflow_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            instance_id TEXT,
            step_id TEXT,
            timestamp TEXT DEFAULT (datetime('now')),
            event_type TEXT,
            agent_id TEXT,
            message TEXT,
            data_json TEXT DEFAULT '{}'
        );

        CREATE INDEX IF NOT EXISTS idx_wf_inst_status ON workflow_instances(status);
        CREATE INDEX IF NOT EXISTS idx_wf_steps_inst ON workflow_steps(instance_id);
        CREATE INDEX IF NOT EXISTS idx_wf_log_inst ON workflow_log(instance_id);
        CREATE INDEX IF NOT EXISTS idx_wf_log_ts ON workflow_log(timestamp);
    """)
    conn.commit()
    return conn


@contextmanager
def _db():
    conn = _init_db()
    try:
        yield conn
        conn.commit()
    finally:
        conn.close()


def _seed_templates():
    """Seed workflow templates into DB (idempotent)."""
    with _db() as conn:
        for tmpl in WORKFLOW_TEMPLATES:
            existing = conn.execute(
                "SELECT id FROM workflow_templates WHERE id=?", (tmpl["id"],)
            ).fetchone()
            if not existing:
                conn.execute(
                    "INSERT INTO workflow_templates (id, name, category, description, icon, steps_json) VALUES (?,?,?,?,?,?)",
                    (tmpl["id"], tmpl["name"], tmpl["category"], tmpl["description"],
                     tmpl.get("icon", ""), json.dumps(tmpl["steps"]))
                )
    logger.info("WorkOps: templates seeded (%d templates)", len(WORKFLOW_TEMPLATES))


# ── Public API ─────────────────────────────────────────────────────────────────

def get_templates() -> List[Dict]:
    """Return all workflow templates with their steps."""
    tmpl_dict = {t["id"]: t for t in WORKFLOW_TEMPLATES}
    with _db() as conn:
        rows = conn.execute("SELECT * FROM workflow_templates ORDER BY category, name").fetchall()
    result = []
    for r in rows:
        t = dict(r)
        t["steps"] = json.loads(t["steps_json"])
        t["icon"] = tmpl_dict.get(t["id"], {}).get("icon", "")
        t["priority_default"] = tmpl_dict.get(t["id"], {}).get("priority_default", "medium")
        del t["steps_json"]
        result.append(t)
    return result


def start_workflow(template_id: str, account_id: str = "system",
                   priority: str = None, context: Dict = None) -> Dict:
    """Start a new workflow instance from a template."""
    tmpl = next((t for t in WORKFLOW_TEMPLATES if t["id"] == template_id), None)
    if not tmpl:
        return {"error": f"Template {template_id!r} not found"}

    instance_id = str(uuid.uuid4())[:12]
    now = datetime.now(timezone.utc).isoformat()
    pri = priority or tmpl.get("priority_default", "medium")

    with _db() as conn:
        conn.execute(
            """INSERT INTO workflow_instances
               (id, template_id, name, category, status, priority, current_step,
                total_steps, started_at, owner_agent, account_id, context_json)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
            (instance_id, template_id, tmpl["name"], tmpl["category"],
             "active", pri, 0, len(tmpl["steps"]), now,
             tmpl["steps"][0]["pickup_agent"] if tmpl["steps"] else "system",
             account_id, json.dumps(context or {}))
        )

        for step in tmpl["steps"]:
            step_id = f"{instance_id}_s{step['index']}"
            conn.execute(
                """INSERT INTO workflow_steps
                   (id, instance_id, step_index, step_name, pickup_agent, putdown_agent, status)
                   VALUES (?,?,?,?,?,?,?)""",
                (step_id, instance_id, step["index"], step["name"],
                 step["pickup_agent"], step["putdown_agent"],
                 "active" if step["index"] == 0 else "pending")
            )

        conn.execute(
            """INSERT INTO workflow_log (instance_id, event_type, agent_id, message)
               VALUES (?,?,?,?)""",
            (instance_id, "workflow_started", "system",
             f"Workflow '{tmpl['name']}' started (priority={pri})")
        )

    logger.info("WorkOps: started workflow %s (%s)", instance_id, tmpl["name"])
    return {"instance_id": instance_id, "name": tmpl["name"], "status": "active", "steps": len(tmpl["steps"])}


def pickup_step(instance_id: str, step_index: int, agent_id: str,
                pickup_data: Dict = None) -> Dict:
    """Agent picks up a step — logs the pickup timestamp and data."""
    now = datetime.now(timezone.utc).isoformat()
    step_id = f"{instance_id}_s{step_index}"

    with _db() as conn:
        existing = conn.execute(
            "SELECT * FROM workflow_steps WHERE id=?", (step_id,)
        ).fetchone()
        if not existing:
            return {"error": f"Step {step_id!r} not found"}

        conn.execute(
            "UPDATE workflow_steps SET pickup_at=?, pickup_data=?, status=? WHERE id=?",
            (now, json.dumps(pickup_data or {}), "active", step_id)
        )
        conn.execute(
            "UPDATE workflow_instances SET current_step=?, owner_agent=? WHERE id=?",
            (step_index, agent_id, instance_id)
        )
        conn.execute(
            """INSERT INTO workflow_log (instance_id, step_id, event_type, agent_id, message, data_json)
               VALUES (?,?,?,?,?,?)""",
            (instance_id, step_id, "step_pickup", agent_id,
             f"Step {step_index} picked up by {agent_id}",
             json.dumps(pickup_data or {}))
        )

    return {"status": "picked_up", "step_id": step_id, "agent_id": agent_id, "pickup_at": now}


def putdown_step(instance_id: str, step_index: int, agent_id: str,
                 putdown_data: Dict = None, handoff_notes: str = "",
                 next_agent_id: str = None) -> Dict:
    """Agent puts down a step — logs completion, auto-advances to next step."""
    now = datetime.now(timezone.utc).isoformat()
    step_id = f"{instance_id}_s{step_index}"

    with _db() as conn:
        step = conn.execute(
            "SELECT * FROM workflow_steps WHERE id=?", (step_id,)
        ).fetchone()
        if not step:
            return {"error": f"Step {step_id!r} not found"}
        _step_name_for_roi = step["step_name"]
        _wf_row = conn.execute("SELECT name FROM workflow_instances WHERE id=?", (instance_id,)).fetchone()
        _wf_name_for_roi = _wf_row["name"] if _wf_row else instance_id

        pickup_at = step["pickup_at"]
        duration_s = None
        if pickup_at:
            try:
                from datetime import datetime as dt
                t0 = dt.fromisoformat(pickup_at.replace("Z", "+00:00"))
                t1 = dt.fromisoformat(now.replace("Z", "+00:00"))
                duration_s = (t1 - t0).total_seconds()
            except Exception:
                pass

        conn.execute(
            """UPDATE workflow_steps
               SET putdown_at=?, putdown_data=?, handoff_notes=?, status=?, duration_s=?
               WHERE id=?""",
            (now, json.dumps(putdown_data or {}), handoff_notes, "complete", duration_s, step_id)
        )
        conn.execute(
            """INSERT INTO workflow_log (instance_id, step_id, event_type, agent_id, message, data_json)
               VALUES (?,?,?,?,?,?)""",
            (instance_id, step_id, "step_putdown", agent_id,
             f"Step {step_index} completed by {agent_id} (dur={duration_s:.0f}s)" if duration_s else f"Step {step_index} completed by {agent_id}",
             json.dumps(putdown_data or {}))
        )

        # Check if there's a next step
        next_step_row = conn.execute(
            "SELECT * FROM workflow_steps WHERE instance_id=? AND step_index=?",
            (instance_id, step_index + 1)
        ).fetchone()

        if next_step_row:
            next_id = f"{instance_id}_s{step_index + 1}"
            conn.execute(
                "UPDATE workflow_steps SET status=? WHERE id=?",
                ("active", next_id)
            )
            next_agent = next_agent_id or next_step_row["pickup_agent"]
            conn.execute(
                "UPDATE workflow_instances SET current_step=?, owner_agent=? WHERE id=?",
                (step_index + 1, next_agent, instance_id)
            )
            conn.execute(
                """INSERT INTO workflow_log (instance_id, step_id, event_type, agent_id, message)
                   VALUES (?,?,?,?,?)""",
                (instance_id, next_id, "step_handoff", agent_id,
                 f"Handed off from {agent_id} → {next_agent} at step {step_index + 1}")
            )
            # Auto-log to ROI ledger
            try:
                from src.roi_ledger import log_workflow_step_complete as _roi_log
                _roi_log(instance_id, _wf_name_for_roi, _step_name_for_roi, agent_id,
                         duration_s or 0, putdown_data)
            except Exception:
                pass
            return {"status": "handed_off", "next_step": step_index + 1, "next_agent": next_agent, "duration_s": duration_s}
        else:
            # Last step — complete the workflow
            conn.execute(
                "UPDATE workflow_instances SET status=?, completed_at=? WHERE id=?",
                ("complete", now, instance_id)
            )
            conn.execute(
                """INSERT INTO workflow_log (instance_id, event_type, agent_id, message)
                   VALUES (?,?,?,?)""",
                (instance_id, "workflow_complete", agent_id,
                 f"Workflow completed — all {step_index + 1} steps done")
            )
            # Auto-log to ROI ledger
        try:
            from src.roi_ledger import log_workflow_step_complete as _roi_log
            _roi_log(instance_id, _wf_name_for_roi, _step_name_for_roi, agent_id,
                     duration_s or 0, putdown_data)
        except Exception:
            pass
        return {"status": "workflow_complete", "completed_at": now, "duration_s": duration_s}


def get_instances(status: str = None, limit: int = 50) -> List[Dict]:
    """List workflow instances with step progress."""
    with _db() as conn:
        if status:
            rows = conn.execute(
                "SELECT * FROM workflow_instances WHERE status=? ORDER BY created_at DESC LIMIT ?",
                (status, limit)
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM workflow_instances ORDER BY created_at DESC LIMIT ?",
                (limit,)
            ).fetchall()

        result = []
        for r in rows:
            inst = dict(r)
            steps = conn.execute(
                "SELECT * FROM workflow_steps WHERE instance_id=? ORDER BY step_index",
                (r["id"],)
            ).fetchall()
            inst["steps"] = [dict(s) for s in steps]
            inst["progress_pct"] = round(
                sum(1 for s in inst["steps"] if s["status"] == "complete") /
                max(len(inst["steps"]), 1) * 100, 0
            )
            result.append(inst)
    return result


def get_orgchart_overlay() -> Dict:
    """Return org chart data with active workflow overlay."""
    # Org structure from REPORT_TO_CHAIN
    org = {
        "rosetta":    {"name": "Rosetta", "emoji": "🌐", "role": "Soul & Constitution", "reports_to": None, "level": 0},
        "exec_admin": {"name": "ExecAdmin", "emoji": "👔", "role": "Executive Director", "reports_to": "rosetta", "level": 1},
        "auditor":    {"name": "Auditor", "emoji": "📋", "role": "Compliance & Audit", "reports_to": "rosetta", "level": 1},
        "hitl":       {"name": "HITL Gate", "emoji": "🔴", "role": "Human-in-the-Loop", "reports_to": "rosetta", "level": 1},
        "translator": {"name": "Translator", "emoji": "🧠", "role": "Intent & NL Processing", "reports_to": "exec_admin", "level": 2},
        "scheduler":  {"name": "Scheduler", "emoji": "🗓", "role": "Task Scheduling", "reports_to": "exec_admin", "level": 2},
        "executor":   {"name": "Executor", "emoji": "⚡", "role": "Action Execution", "reports_to": "exec_admin", "level": 2},
        "prod_ops":   {"name": "ProdOps", "emoji": "🔧", "role": "Production Operations", "reports_to": "exec_admin", "level": 2},
        "collector":  {"name": "Collector", "emoji": "📡", "role": "Signal Collection", "reports_to": "translator", "level": 3},
    }

    with _db() as conn:
        # Get active workflows and which agent owns each
        active = conn.execute(
            "SELECT owner_agent, COUNT(*) as c, GROUP_CONCAT(name) as workflows "
            "FROM workflow_instances WHERE status='active' GROUP BY owner_agent"
        ).fetchall()

        for row in active:
            agent_id = row["owner_agent"]
            if agent_id in org:
                org[agent_id]["active_workflows"] = row["c"]
                org[agent_id]["workflow_names"] = row["workflows"]

        # Recent log events per agent
        recent = conn.execute(
            """SELECT agent_id, COUNT(*) as events, MAX(timestamp) as last_active
               FROM workflow_log WHERE timestamp > datetime('now', '-24 hours')
               GROUP BY agent_id"""
        ).fetchall()
        for row in recent:
            if row["agent_id"] in org:
                org[row["agent_id"]]["events_24h"] = row["events"]
                org[row["agent_id"]]["last_active"] = row["last_active"]

        # Stats
        stats = conn.execute(
            """SELECT
                SUM(CASE WHEN status='active' THEN 1 ELSE 0 END) as active_count,
                SUM(CASE WHEN status='complete' THEN 1 ELSE 0 END) as complete_count,
                SUM(CASE WHEN status='pending' THEN 1 ELSE 0 END) as pending_count
               FROM workflow_instances"""
        ).fetchone()

    return {
        "agents": org,
        "stats": dict(stats) if stats else {},
        "templates": [{"id": t["id"], "name": t["name"], "category": t["category"], "icon": t.get("icon", "")}
                      for t in WORKFLOW_TEMPLATES]
    }


def get_recent_log(limit: int = 100) -> List[Dict]:
    with _db() as conn:
        rows = conn.execute(
            "SELECT * FROM workflow_log ORDER BY timestamp DESC LIMIT ?", (limit,)
        ).fetchall()
    return [dict(r) for r in rows]


def export_excel() -> bytes:
    """Generate Excel workbook with 4 tabs: Operations Log, Org View, Business Plan vs Active, KPIs."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        import subprocess
        subprocess.run(["pip", "install", "openpyxl", "-q"], check=True)
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ── Styles ────────────────────────────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor="0D2137")
    header_font = Font(bold=True, color="00D4FF", size=10)
    alt_fill    = PatternFill("solid", fgColor="0A1628")
    green_fill  = PatternFill("solid", fgColor="064420")
    amber_fill  = PatternFill("solid", fgColor="3D2800")
    red_fill    = PatternFill("solid", fgColor="3D0A0A")
    title_font  = Font(bold=True, color="00FF88", size=12)
    cell_font   = Font(color="C8D8E8", size=9)
    thin        = Border(
        left=Side(style="thin", color="1E2D45"),
        right=Side(style="thin", color="1E2D45"),
        top=Side(style="thin", color="1E2D45"),
        bottom=Side(style="thin", color="1E2D45")
    )

    def style_header_row(ws, row, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin

    def style_data_row(ws, row, cols, fill=None):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = cell_font
            cell.border = thin
            if fill:
                cell.fill = fill

    with _db() as conn:
        instances = conn.execute("SELECT * FROM workflow_instances ORDER BY created_at DESC LIMIT 500").fetchall()
        steps_all = conn.execute("SELECT * FROM workflow_steps ORDER BY created_at DESC LIMIT 2000").fetchall()
        log_rows  = conn.execute("SELECT * FROM workflow_log ORDER BY timestamp DESC LIMIT 1000").fetchall()

    # ── Tab 1: Operations Log ─────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Operations Log"
    ws1.sheet_view.showGridLines = False
    ws1.freeze_panes = "A3"

    ws1.merge_cells("A1:K1")
    ws1["A1"] = "MURPHY SYSTEM — OPERATIONS LOG"
    ws1["A1"].font = title_font
    ws1["A1"].fill = PatternFill("solid", fgColor="061020")
    ws1["A1"].alignment = Alignment(horizontal="center")
    ws1.row_dimensions[1].height = 24

    headers1 = ["Workflow ID", "Workflow Name", "Category", "Status", "Priority",
                "Step #", "Step Name", "Pickup Agent", "Pickup At",
                "Putdown Agent", "Putdown At", "Duration (s)", "Handoff Notes"]
    for c, h in enumerate(headers1, 1):
        ws1.cell(row=2, column=c).value = h
    style_header_row(ws1, 2, len(headers1))
    ws1.row_dimensions[2].height = 20

    inst_dict = {r["id"]: dict(r) for r in instances}
    row_n = 3
    for s in steps_all:
        inst = inst_dict.get(s["instance_id"], {})
        status = s["status"]
        fill = green_fill if status == "complete" else (amber_fill if status == "active" else None)
        vals = [
            s["instance_id"], inst.get("name", ""), inst.get("category", ""),
            inst.get("status", ""), inst.get("priority", ""),
            s["step_index"], s["step_name"],
            s["pickup_agent"], s["pickup_at"] or "",
            s["putdown_agent"], s["putdown_at"] or "",
            round(s["duration_s"], 1) if s["duration_s"] else "",
            s["handoff_notes"] or ""
        ]
        for c, v in enumerate(vals, 1):
            cell = ws1.cell(row=row_n, column=c, value=v)
            cell.font = cell_font
            cell.border = thin
            if fill:
                cell.fill = fill
        row_n += 1

    col_widths1 = [14, 22, 16, 10, 10, 7, 20, 14, 20, 14, 20, 12, 30]
    for i, w in enumerate(col_widths1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ── Tab 2: Org Structure ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("Org Structure")
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = "A3"

    ws2.merge_cells("A1:H1")
    ws2["A1"] = "MURPHY SWARM — ORG STRUCTURE & LIVE LOAD"
    ws2["A1"].font = title_font
    ws2["A1"].fill = PatternFill("solid", fgColor="061020")
    ws2["A1"].alignment = Alignment(horizontal="center")
    ws2.row_dimensions[1].height = 24

    org_data = get_orgchart_overlay()
    headers2 = ["Agent ID", "Name", "Emoji", "Role", "Reports To", "Level",
                "Active Workflows", "Events (24h)", "Last Active"]
    for c, h in enumerate(headers2, 1):
        ws2.cell(row=2, column=c).value = h
    style_header_row(ws2, 2, len(headers2))

    level_fills = {0: PatternFill("solid", fgColor="1A0A3A"),
                   1: PatternFill("solid", fgColor="0A1A2A"),
                   2: PatternFill("solid", fgColor="061420"),
                   3: PatternFill("solid", fgColor="04100A")}
    row_n = 3
    for agent_id, info in sorted(org_data["agents"].items(), key=lambda x: x[1]["level"]):
        vals = [
            agent_id, info["name"], info["emoji"], info["role"],
            info["reports_to"] or "(top)", info["level"],
            info.get("active_workflows", 0),
            info.get("events_24h", 0),
            info.get("last_active", "")
        ]
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(row=row_n, column=c, value=v)
            cell.font = cell_font
            cell.border = thin
            cell.fill = level_fills.get(info["level"], alt_fill)
        row_n += 1

    col_widths2 = [14, 14, 8, 24, 14, 8, 18, 14, 24]
    for i, w in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── Tab 3: Business Plan vs Active ────────────────────────────────────────
    ws3 = wb.create_sheet("Business Plan vs Active")
    ws3.sheet_view.showGridLines = False
    ws3.freeze_panes = "A3"

    ws3.merge_cells("A1:L1")
    ws3["A1"] = "MURPHY SYSTEM — BUSINESS PLAN (TEMPLATE) vs ACTIVE OPERATIONS"
    ws3["A1"].font = title_font
    ws3["A1"].fill = PatternFill("solid", fgColor="061020")
    ws3["A1"].alignment = Alignment(horizontal="center")
    ws3.row_dimensions[1].height = 24

    headers3 = ["Category", "Workflow Name", "Step #", "Step Name",
                "Planned Pickup Agent", "Planned Duration (m)", "Pickup Form Fields",
                "Active Instance ID", "Instance Status", "Actual Pickup Agent",
                "Actual Duration (s)", "Variance (s)", "Handoff Notes"]
    for c, h in enumerate(headers3, 1):
        ws3.cell(row=2, column=c).value = h
    style_header_row(ws3, 2, len(headers3))

    row_n = 3
    for tmpl in WORKFLOW_TEMPLATES:
        # Find most recent active instance of this template
        active_inst = None
        active_steps = {}
        with _db() as conn:
            ai = conn.execute(
                "SELECT * FROM workflow_instances WHERE template_id=? ORDER BY created_at DESC LIMIT 1",
                (tmpl["id"],)
            ).fetchone()
            if ai:
                active_inst = dict(ai)
                step_rows = conn.execute(
                    "SELECT * FROM workflow_steps WHERE instance_id=?",
                    (ai["id"],)
                ).fetchall()
                active_steps = {s["step_index"]: dict(s) for s in step_rows}

        cat_fill = PatternFill("solid", fgColor="0A1628")
        for step in tmpl["steps"]:
            aStep = active_steps.get(step["index"], {})
            planned_dur = step.get("expected_duration_m", 0) * 60
            actual_dur  = aStep.get("duration_s") or 0
            variance    = round(actual_dur - planned_dur, 0) if actual_dur and planned_dur else ""
            var_fill    = (green_fill if isinstance(variance, (int, float)) and variance < 0
                           else red_fill if isinstance(variance, (int, float)) and variance > 120
                           else cat_fill)
            vals = [
                tmpl["category"], tmpl["name"], step["index"], step["name"],
                step["pickup_agent"],
                step.get("expected_duration_m", ""),
                ", ".join(step.get("pickup_form", [])),
                active_inst["id"] if active_inst else "",
                active_inst["status"] if active_inst else "no instance",
                aStep.get("pickup_agent", ""),
                round(actual_dur, 0) if actual_dur else "",
                variance,
                aStep.get("handoff_notes", "")
            ]
            for c, v in enumerate(vals, 1):
                cell = ws3.cell(row=row_n, column=c, value=v)
                cell.font = cell_font
                cell.border = thin
                cell.fill = var_fill if c >= 12 else cat_fill
            row_n += 1

    col_widths3 = [18, 22, 7, 22, 18, 18, 32, 16, 14, 18, 14, 12, 30]
    for i, w in enumerate(col_widths3, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # ── Tab 4: KPIs ───────────────────────────────────────────────────────────
    ws4 = wb.create_sheet("KPIs")
    ws4.sheet_view.showGridLines = False

    ws4.merge_cells("A1:E1")
    ws4["A1"] = "MURPHY SYSTEM — WORKFLOW KPIs"
    ws4["A1"].font = title_font
    ws4["A1"].fill = PatternFill("solid", fgColor="061020")
    ws4["A1"].alignment = Alignment(horizontal="center")
    ws4.row_dimensions[1].height = 24

    kpi_headers = ["Metric", "Value", "Unit", "Notes"]
    for c, h in enumerate(kpi_headers, 1):
        ws4.cell(row=2, column=c).value = h
    style_header_row(ws4, 2, len(kpi_headers))

    with _db() as conn:
        total_inst  = conn.execute("SELECT COUNT(*) FROM workflow_instances").fetchone()[0]
        active_inst_count = conn.execute("SELECT COUNT(*) FROM workflow_instances WHERE status='active'").fetchone()[0]
        complete_count = conn.execute("SELECT COUNT(*) FROM workflow_instances WHERE status='complete'").fetchone()[0]
        avg_dur     = conn.execute("SELECT AVG((julianday(completed_at)-julianday(started_at))*86400) FROM workflow_instances WHERE status='complete'").fetchone()[0]
        steps_total = conn.execute("SELECT COUNT(*) FROM workflow_steps").fetchone()[0]
        steps_done  = conn.execute("SELECT COUNT(*) FROM workflow_steps WHERE status='complete'").fetchone()[0]
        avg_step_dur = conn.execute("SELECT AVG(duration_s) FROM workflow_steps WHERE duration_s IS NOT NULL").fetchone()[0]
        bottleneck  = conn.execute(
            "SELECT pickup_agent, AVG(duration_s) as avg FROM workflow_steps WHERE duration_s IS NOT NULL GROUP BY pickup_agent ORDER BY avg DESC LIMIT 1"
        ).fetchone()

    kpis = [
        ("Total Workflow Instances", total_inst, "count", "All time"),
        ("Active Instances", active_inst_count, "count", "Currently running"),
        ("Completed Instances", complete_count, "count", "All time"),
        ("Completion Rate", f"{round(complete_count/max(total_inst,1)*100,1)}%", "%", "complete/total"),
        ("Avg Workflow Duration", round(avg_dur, 0) if avg_dur else "N/A", "seconds", "Completed only"),
        ("Total Steps Logged", steps_total, "count", "All instances"),
        ("Steps Completed", steps_done, "count", "All instances"),
        ("Avg Step Duration", round(avg_step_dur, 1) if avg_step_dur else "N/A", "seconds", "Across all agents"),
        ("Slowest Agent", bottleneck["pickup_agent"] if bottleneck else "N/A", "agent_id",
         f"{round(bottleneck['avg'],0):.0f}s avg" if bottleneck and bottleneck['avg'] else ""),
        ("Workflow Templates", len(WORKFLOW_TEMPLATES), "count", "9 standard workflows"),
    ]

    for i, (metric, value, unit, notes) in enumerate(kpis, 3):
        row_fill = alt_fill if i % 2 == 0 else PatternFill("solid", fgColor="04100A")
        ws4.cell(row=i, column=1, value=metric).fill = row_fill
        ws4.cell(row=i, column=2, value=value).fill = row_fill
        ws4.cell(row=i, column=3, value=unit).fill = row_fill
        ws4.cell(row=i, column=4, value=notes).fill = row_fill
        for c in range(1, 5):
            ws4.cell(row=i, column=c).font = cell_font
            ws4.cell(row=i, column=c).border = thin

    for w, cw in zip("ABCD", [30, 16, 12, 30]):
        ws4.column_dimensions[w].width = cw

    # ── Write to bytes ─────────────────────────────────────────────────────────
    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Init on import ─────────────────────────────────────────────────────────────
try:
    _seed_templates()
except Exception as e:
    logger.warning("WorkOps seed error: %s", e)
